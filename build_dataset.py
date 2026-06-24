import openpyxl, json, re, sys

SRC = "/sessions/gallant-affectionate-bell/mnt/Chat Derivaciones/RE-020 Registros HEXALIS (Planillas paridades Datatech- Hexalis).xlsx"

def norm(s):
    return re.sub(r'\s+', ' ', str(s)).strip().upper() if s is not None else ''

FIELD_PATTERNS = {
    'codigo_hexalis': ['CODIGO HEXALIS'],
    'nemonico_datatech': ['NEMONICO DATATECH'],
    'paridad': ['PARIDAD DATATECH'],
    'estudio': ['ESTUDIO', 'ANALISIS', 'ESTUDIOS'],
    'metodo': ['METODO'],
    'resultado_preanalitico': ['RESULTADO PREANALITICO'],
    'muestra': ['MUESTRA'],
    'indicaciones': ['INDICACIONES'],
    'valores_referencia': ['VALORES DE REFERENCIA', 'VALOR DE REFERENCIA'],
    'demora': ['DEMORA'],
    'status': ['STATUS'],
    'observaciones': ['OBSERVACIONES'],
    'modelo_informe': ['MODELO DE INFORME'],
    'reviso': ['REVISO'],
    'nota': ['NOTA'],
    'resultado': ['RESULTADO', 'RESULTADOS'],
}

def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        vals = [norm(c) for c in row]
        if any(v == 'METODO' for v in vals) and any('MUESTRA' == v for v in vals):
            return i, row
    return None, None

def map_columns(headers):
    mapping = {}
    used = set()
    # pass 1: exact match
    for idx, h in enumerate(headers):
        hn = norm(h)
        if not hn:
            continue
        for field, patterns in FIELD_PATTERNS.items():
            if field in mapping:
                continue
            if hn in patterns:
                mapping[field] = idx
                used.add(idx)
                break
    # pass 2: fuzzy substring match for remaining fields, skipping used columns
    for idx, h in enumerate(headers):
        if idx in used:
            continue
        hn = norm(h)
        if not hn:
            continue
        for field, patterns in FIELD_PATTERNS.items():
            if field in mapping:
                continue
            for p in patterns:
                if p in hn:
                    mapping[field] = idx
                    used.add(idx)
                    break
    return mapping

wb = openpyxl.load_workbook(SRC, data_only=True)
records = []
for ws in wb.worksheets:
    if ws.title == 'ORDEN':
        continue
    hr, headers = find_header_row(ws)
    if hr is None:
        print('NO HEADER FOUND for', ws.title, file=sys.stderr)
        continue
    colmap = map_columns(headers)
    for row in ws.iter_rows(min_row=hr+1, values_only=True):
        if not any(c is not None for c in row):
            continue
        rec = {'laboratorio': ws.title}
        for field, idx in colmap.items():
            if idx < len(row):
                val = row[idx]
                if val is not None:
                    rec[field] = str(val).strip() if not isinstance(val, (int, float)) else val
        # skip rows without a study name
        if not rec.get('estudio'):
            continue
        records.append(rec)

print(f'Total records: {len(records)}', file=sys.stderr)
with open('/sessions/gallant-affectionate-bell/mnt/outputs/chatbot/estudios.json', 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False, indent=1)
